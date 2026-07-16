import sys
import json
import os

try:
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Border, Side, Alignment, PatternFill
except ImportError:
    # MVP fallback if openpyxl is not installed natively
    print(json.dumps({"status": "error", "message": "openpyxl not installed"}))
    sys.exit(1)

def generate_excel(payload_json):
    try:
        data = json.loads(payload_json)
        
        # Mapeo de datos básicos
        opp_id = data.get('opportunityId', 'UNKNOWN')
        property_data = data.get('property', {})
        matrix_data = data.get('matrix', '0')
        photos = data.get('photos', [])
        nombre_proyecto = str(property_data.get('nombreEdificio', '')).upper()
        
        template_name = "Ficha de Datos NOMBRE DEL PROYECTO - Junio.xlsx"
        template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', template_name)
        
        # Resolve to absolute path to be sure
        template_path = os.path.abspath(template_path)
        
        # Simulación de carga de plantilla corporativa real
        if os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
        else:
            # Fallback en caso de que la plantilla física no esté subida aún al contenedor
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ficha Tecnica"
        
        def safe_write(ws, cell, val):
            try:
                ws[cell] = val
            except Exception:
                pass

        # Inyectando campos validados en MAYÚSCULAS
        safe_write(ws, 'C5', nombre_proyecto) # NOMBRE DEL EDIFICIO/CONDOMINIO (C5:M5)
        
        # 1. TIPO DE PROYECTO (F8: NUEVO PREDIO, L8: AMPLIACION DE TORRE)
        tipo_desarrollo = str(property_data.get('tipoDesarrollo', '')).upper()
        if 'AMPLIACION' in tipo_desarrollo or 'CONDOMINIO' in tipo_desarrollo:
            safe_write(ws, 'L8', 'X')
        else:
            safe_write(ws, 'F8', 'X')
            
        # 2. FUENTE/ORIGEN (E12: PROPIO, M12: REFERIDO)
        origen = str(property_data.get('origen', '')).upper()
        if 'REFERIDO' in origen:
            safe_write(ws, 'M12', 'X')
            safe_write(ws, 'E12', '')
        else:
            safe_write(ws, 'E12', 'X')
            safe_write(ws, 'M12', '')
            
        # 3. CLASIFICACION (E16: EDIFICIO, I16: CONDOMINIO)
        clasificacion = str(property_data.get('clasificacion', '')).upper()
        total_torres = int(property_data.get('totalTorres', 1))
        if 'CONDOMINIO' in clasificacion or total_torres >= 3:
            safe_write(ws, 'I16', 'X')
            safe_write(ws, 'E16', '')
        else:
            safe_write(ws, 'E16', 'X')
            safe_write(ws, 'I16', '')
            
        # 4. TIPO DE CONSTRUCCION (E22: ESTRENO, I22: MODERNO, L22: ANTIGUO)
        tipo_construccion = str(property_data.get('tipoConstruccion', '')).upper()
        if 'ESTRENO' in tipo_construccion:
            safe_write(ws, 'E22', 'X')
            safe_write(ws, 'I22', '')
            safe_write(ws, 'L22', '')
            # Si es Estreno, inyectamos las fechas correspondientes
            safe_write(ws, 'E23', str(property_data.get('fechaEntrega', '')))
            safe_write(ws, 'E25', str(property_data.get('fechaMontantes', '')))
        elif 'ANTIGUO' in tipo_construccion:
            safe_write(ws, 'L22', 'X')
            safe_write(ws, 'E22', '')
            safe_write(ws, 'I22', '')
        else:
            safe_write(ws, 'I22', 'X')
            safe_write(ws, 'E22', '')
            safe_write(ws, 'L22', '')

        # JUNTA DIRECTIVA (E31: SI, I31: NO)
        junta_directiva = str(property_data.get('juntaDirectiva', '')).upper()
        if 'SI' in junta_directiva or 'SÍ' in junta_directiva:
            safe_write(ws, 'E31', 'X')
            safe_write(ws, 'I31', '')
        else:
            safe_write(ws, 'I31', 'X')
            safe_write(ws, 'E31', '')

        # RESPONSABLE
        safe_write(ws, 'D34', str(property_data.get('cargoResponsable', '')).upper()) # CARGO
        safe_write(ws, 'I34', str(property_data.get('responsable', '')).upper()) # NOMBRE
        safe_write(ws, 'D35', str(property_data.get('telefonoResponsable', ''))) # CELULAR
        safe_write(ws, 'I35', str(property_data.get('correoResponsable', '')).upper()) # CORREO

        # VISITA TECNICA
        safe_write(ws, 'D39', str(property_data.get('fechaVisitaTecnica', ''))) # FECHA
        horario = str(property_data.get('horarioVisita', '')).upper()
        if '9' in horario or '12' in horario or ('AM' in horario and '1' not in horario):
            safe_write(ws, 'I39', 'X')
            safe_write(ws, 'L39', '')
        elif '1' in horario or '4' in horario or 'PM' in horario:
            safe_write(ws, 'L39', 'X')
            safe_write(ws, 'I39', '')
        else:
            safe_write(ws, 'I39', '')
            safe_write(ws, 'L39', '')

        # 5. DIRECCION Y COORDENADAS
        safe_write(ws, 'C43', str(property_data.get('departamento', '')).upper()) # DEPARTAMENTO
        safe_write(ws, 'I43', str(property_data.get('provincia', '')).upper()) # PROVINCIA
        safe_write(ws, 'C44', str(property_data.get('distrito', '')).upper()) # DISTRITO
        safe_write(ws, 'I44', str(property_data.get('urbanizacion', '')).upper()) # URBANIZACION
        safe_write(ws, 'C45', str(property_data.get('codigoPostal', ''))) # CODIGO POSTAL
        safe_write(ws, 'C46', str(property_data.get('tipoVia', '')).upper()) # TIPO VIA
        safe_write(ws, 'C47', str(property_data.get('nombreVia', '')).upper()) # NOMBRE VIA
        safe_write(ws, 'C48', str(property_data.get('numeracion', '')).upper()) # NUMERO VIA
        safe_write(ws, 'C49', str(property_data.get('coordenadas', ''))) # COORDENADAS
        
        safe_write(ws, 'C53', total_torres) # TOTAL TORRES
        safe_write(ws, 'C54', int(property_data.get('totalHogares', 0))) # TOTAL HOGARES
            
        # Matriz de Torres Relacional
        matrix_list = property_data.get('matrixList', []) if isinstance(property_data, dict) else []
        if not matrix_list:
            matrix_list = data.get('matrixList', [])
        if not matrix_list:
            matrix_list = [matrix_data]
            
        col_letters = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        bloques_filas = [58, 61, 64, 67]
        bloque_idx = 0
        
        towers = []
        for idx, tower_matrix in enumerate(matrix_list):
            if idx >= 4: break
            pisos = [p.strip() for p in str(tower_matrix).split(',') if p.strip()]
            if not pisos: continue
            
            hogares_list = []
            for p in pisos:
                try:
                    hogares_list.append(int(p))
                except ValueError:
                    hogares_list.append(0)
            
            towers.append({
                "name": str(idx + 1),
                "pisos": len(pisos),
                "hogares_list": hogares_list
            })
            
        for t in towers:
            pisos_restantes = t["pisos"]
            hogares_restantes = list(t["hogares_list"])
            tower_block_num = 0

            while pisos_restantes > 0 and bloque_idx < len(bloques_filas):
                row_start = bloques_filas[bloque_idx]

                if tower_block_num == 0:
                    safe_write(ws, f"A{row_start}", f"TORRE {t['name']}")
                else:
                    safe_write(ws, f"A{row_start}", "")

                safe_write(ws, f"B{row_start}", "PISO :")
                safe_write(ws, f"B{row_start+1}", "HOGARES POR PISO")

                for c in range(10):
                    col_letter = col_letters[c]
                    floor_in_block = tower_block_num * 10 + (c + 1)

                    safe_write(ws, f"{col_letter}{row_start}", floor_in_block)

                    if floor_in_block <= t["pisos"]:
                        h_val = hogares_restantes[floor_in_block - 1]
                        safe_write(ws, f"{col_letter}{row_start+1}", h_val)
                    else:
                        safe_write(ws, f"{col_letter}{row_start+1}", "N/A")

                safe_write(ws, f"M{row_start}", "TOTAL")
                ws[f"M{row_start+1}"] = f"=SUM(C{row_start+1}:L{row_start+1})"

                pisos_restantes -= 10
                tower_block_num += 1
                bloque_idx += 1
                
        # 6. CANAL / AGENCIA
        safe_write(ws, 'C72', str(data.get('canalHunting', '')).upper())
        
        # 7. GESTOR Y SUPERVISOR
        is_referral = data.get('isReferral', False)
        if is_referral:
            safe_write(ws, 'C75', str(data.get('referredHunterName', '')).upper()) # NOMBRE DEL HUNTER
            safe_write(ws, 'I75', 'N/A') # CELULAR DEL HUNTER
        else:
            safe_write(ws, 'C75', str(data.get('currentOwnerName', '')).upper()) # NOMBRE DEL HUNTER
            safe_write(ws, 'I75', str(data.get('currentOwnerPhone', ''))) # CELULAR DEL HUNTER

        # Incrustar imágenes físicamente (estáticamente en A80 y F80 como en el script original)
        cells = ['A80', 'F80']
        for idx, photo_path in enumerate(photos):
            if idx >= len(cells): break
            cell = cells[idx]
            if os.path.exists(photo_path):
                try:
                    from PIL import Image as PILImage
                    final_path = photo_path
                    
                    with PILImage.open(photo_path) as im:
                        real_format = str(im.format).upper()
                        if real_format not in ['PNG', 'JPEG']:
                            png_path = os.path.splitext(photo_path)[0] + '_converted.png'
                            im.convert('RGB').save(png_path, 'PNG')
                            final_path = png_path
                    
                    img = Image(final_path)
                    img.width = 330
                    img.height = 260
                    ws.add_image(img, cell)
                except Exception as e:
                    safe_write(ws, cell, f"Error cargando imagen: {str(e)}")
            else:
                safe_write(ws, cell, "Imagen no encontrada")

        # Guardado final
        output_name = f"Ficha_de_Datos_{nombre_proyecto.replace(' ', '_')}_{opp_id}.xlsx"
        output_path = os.path.join(os.path.dirname(__file__), output_name)
        wb.save(output_path)
        
        print(json.dumps({
            "status": "success",
            "message": "Report generated successfully",
            "file": output_path
        }))
        
    except Exception as e:
        print(json.dumps({"status": "error", "message": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        generate_excel(sys.argv[1])
    else:
        print(json.dumps({"status": "error", "message": "No payload provided"}))
        sys.exit(1)
